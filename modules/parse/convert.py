import io
import os
import csv
import warnings
from datetime import date, datetime, timezone, timedelta
from openpyxl import load_workbook
import xlrd
import pandas as pd
from modules.core.util import clean_date
from modules.ticker import util as tu
from modules.object.provider import Mapping

FILE_FOLDER = "./.downloads/"
DECIMAL_PRECISION = 10

def read_xls_from_buffer(file_buffer: bytes, mapping: Mapping) -> list[list[str]]:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style*")
        book = xlrd.open_workbook(file_contents=file_buffer)
        sheet_name = mapping.sheet if mapping.sheet is not None else book.sheet_names()[0]
        sheet = book.sheet_by_name(sheet_name)
        full_rows = []
        for r in range(sheet.nrows):
            full_rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
        return full_rows
    
def read_xlsx_from_buffer(file_buffer: bytes, mapping: Mapping) -> list[list[str]]:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style*")
        book = load_workbook(io.BytesIO(file_buffer), data_only=True)
        sheet = mapping.sheet if mapping.sheet is not None else book.sheetnames[0]
        ws = book[sheet]
        full_rows = []
        if not ws:
            raise Exception('Could not read raw data from Excel holdings file')
        
        for row in ws.iter_rows(values_only=True):
            full_rows.append(list(row))
        return full_rows
    
def read_csv_from_buffer(file_buffer: str, mapping: Mapping) -> list[list[str]]:
    reader = csv.reader(io.StringIO(file_buffer))
    return list(reader)

def read_file(file_name: str, format: str, mapping: Mapping) -> list[list[str]]:
    if format == 'xls':
        with open(os.path.join(FILE_FOLDER, file_name), "rb") as f:
            data = f.read()
            return read_xls_from_buffer(data, mapping)
    if format == 'xlsx':
        with open(os.path.join(FILE_FOLDER, file_name), "rb") as f:
            data = f.read()
            return read_xlsx_from_buffer(data, mapping)
    else:
        with open(os.path.join(FILE_FOLDER, file_name), "r") as f:
            data = f.read()
            return read_csv_from_buffer(data, mapping)

def detect_single_header_row(
    full_rows: list[list[str]],
    expected_headers: set[str],
    start_row: int = 0,
    max_scan: int = 30,
    min_match_ratio: float = 0.6,
) -> int | None:
    """
    Detect the header row by matching expected column names.
    """
    expected = {h.lower().strip() for h in expected_headers if h}

    for i in range(start_row, min(start_row + max_scan, len(full_rows))):
        row = full_rows[i]
        if not row:
            continue

        normalized = {
            str(cell).lower().strip()
            for cell in row
            if cell not in (None, "", "nan")
        }

        matches = expected & normalized
        if expected and len(matches) / len(expected) >= min_match_ratio:
            return i

    return None

from datetime import datetime
from typing import Optional


def detect_shifted_single_date(
    full_rows: list[list[str]],
    start_row: int,
    col: int,
    date_format: str,
    max_scan: int = 0,
) -> Optional[datetime]:
    """
    Attempt to detect a date in a fixed column by scanning downward
    from a starting row.
    """
    for i in range(start_row, min(start_row + max_scan + 1, len(full_rows))):
        try:
            cell = full_rows[i][col]
        except IndexError:
            continue

        if cell in (None, "", "nan"):
            continue

        text = str(cell).strip()
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue

    return None

def clean_numeric_column(df: pd.DataFrame, column: str, as_type: str = "float") -> pd.Series:
    col = df[column].astype(str).str.strip()
    col_clean = col.str.replace(r"[$€£,%]", "", regex=True)
    col_numeric = pd.to_numeric(col_clean, errors="coerce")
        
    if as_type == "int":
        col_numeric = col_numeric.astype("Int64")  # nullable integer type
    
    return col_numeric


def convert_to_data_frame(full_rows: list[list[str]], mapping: Mapping) -> pd.DataFrame:
    skip = mapping.skip_rows
    if mapping.multi_row_header == 2:
        # Two header rows
        header_1 = full_rows[skip]
        header_2 = full_rows[skip + 1]
        
        # Flatten into a single string
        header_1_fixed = []
        last = ""
        for h in header_1:
            if h not in [None, "", "nan"]:
                last = str(h).strip()
            header_1_fixed.append(last)

        no_prefix = set(getattr(mapping, "no_prefix_headers", []))

        # Now flatten properly using both levels
        header = []
        for top, sub in zip(header_1_fixed, header_2):
            sub = "" if sub is None else str(sub).strip()
            top = str(top).strip()

            if sub in no_prefix or top == "":
                # standalone second-layer header
                header.append(sub)
            else:
                header.append(f"{top} {sub}".strip())

        data_start = skip + mapping.header_data_gap + 2
    else:
        expected_headers = {
            src for src in mapping.columns.values() if src is not None
        }

        detected = detect_single_header_row(full_rows, expected_headers, start_row=mapping.skip_rows)

        skip = detected if detected is not None else mapping.skip_rows

        header = [str(h).strip() if h is not None else "" for h in full_rows[skip]]
        data_start = skip + mapping.header_data_gap + 1

    data = full_rows[data_start:]
    df = pd.DataFrame(data, columns=header)

    df.columns = df.columns.str.replace('\ufeff', '', regex=True)

    return df

def map_data(full_rows: list[list[str]], file_name:str, date_from_page: date | None, mapping: Mapping) -> pd.DataFrame:
    df = convert_to_data_frame(full_rows=full_rows, mapping=mapping)

    good_date: date | None = None
    
    if mapping.date.on_page and date_from_page:
        good_date = date_from_page

    if mapping.date.in_file_name:
        good_date = clean_date(file_name, format=mapping.date.format)

    # If the date is not a part of the table or is only in the first row - grab it:
    if mapping.date.single:
        if mapping.date.single.max_row_scan:
            detected_date = detect_shifted_single_date(
                full_rows,
                start_row=mapping.date.single.row,
                col=mapping.date.single.col,
                date_format=mapping.date.format,
                max_scan=mapping.date.single.max_row_scan
            )

            if detected_date:
                good_date = detected_date
        else:
            dirty = full_rows[mapping.date.single.row][mapping.date.single.col]
            good_date = clean_date(dirty, format=mapping.date.format)

    # select only the rows of the product (if this is a multi-product sheet)
    if mapping.product_column:
        df = df[df[mapping.product_column] == mapping.product_symbol]

    # Select only columns that exist in the original dataframe
    inverse_map = {src.lower(): tgt for tgt, src in mapping.columns.items() if src is not None}
    df.columns = df.columns.str.lower()
    df = df[list(inverse_map.keys())]
    df.rename(columns=lambda c: inverse_map.get(c.lower(), c), inplace=True)

    # Ensure ALL target columns exist, including those mapped from None
    for tgt, src in mapping.columns.items():
        if tgt not in df.columns:
            df.loc[:, tgt] = None

    # Maintain the exact order defined in mapping
    df = df[list(mapping.columns.keys())]
    df = df.dropna(how="all")

    if good_date:
        df.loc[:, "holding_date"] = good_date
    else:
        df["holding_date"] = pd.to_datetime(df["holding_date"], format=mapping.date.format, errors="coerce")
    
    missing_mask = df['ticker'].isna() | (df['ticker'].astype(str).str.strip() == '')
    if missing_mask.any():
        has_isin = 'isin' in df.columns
        for idx in df[missing_mask].index:
            isin = str(df.at[idx, 'isin']) if has_isin and pd.notna(df.at[idx, 'isin']) else None
            name = str(df.at[idx, 'name']) if pd.notna(df.at[idx, 'name']) else None
            symbol = tu.resolve_ticker_from_alt_data(isin=isin, name=name)
            if symbol:
                df.at[idx, 'ticker'] = symbol

    required_cols = ["holding_date", "ticker", "weight", "name"]
    df = df.dropna(subset=required_cols)

    df["market_value"] = clean_numeric_column(df, "market_value")
    df["weight"] = clean_numeric_column(df, "weight")
    df["shares"] = clean_numeric_column(df, "shares")
    if "price" in df.columns:
        df["price"] = clean_numeric_column(df, "price")

    # Fall back to shares × price only when no market_value column is mapped
    has_market_value_mapping = mapping.columns.get("market_value") is not None
    if not has_market_value_mapping and "price" in df.columns:
        df["market_value"] = df["shares"] * df["price"]

    # Scale market_value if the source reports in thousands, millions, etc.
    if mapping.market_value.shift:
        df["market_value"] = df["market_value"] * (10 ** mapping.market_value.shift)

    df = df[df.apply(lambda row: tu.is_valid_holding(row['ticker'], row.get('name')), axis=1)]
    df['ticker'] = tu.normalize_ticker_series(df['ticker'])
    df = tu.filter_ticker_df(df, 'ticker', mapping.remove_tickers)

    # drop rows where the shares, market_value or weight is empty or 0 (used for cash holdings)
    df = df[df["shares"].notna() & (df["shares"] > 0)]
    df = df[df["market_value"].notna() & (df["market_value"] > 0)]
    df = df[df["weight"].notna() & (df["weight"] > 0)]

    # We are using a weighting that sums up to 1 (not 100%)
    total_weight = df["weight"].sum()
    if total_weight > 1:
        df["weight"] = df["weight"] / 100
        df["weight"] = df["weight"].round(DECIMAL_PRECISION)

    df = df.drop(columns=["price"], errors="ignore")

    # Sum up holdings that have the same ticker; take latest holding_date if rows differ
    agg: dict = {
        'market_value': 'sum',
        'shares': 'sum',
        'weight': 'sum',
        'holding_date': 'max',
    }
    for col in ('isin', 'cusip', 'sedol', 'name'):
        if col in df.columns:
            agg[col] = 'first'
    df = df.groupby('ticker', as_index=False).agg(agg)

    return df    

def get_tickers(full_rows: list[list[str]], mapping: Mapping) -> list[str]:
    df = convert_to_data_frame(full_rows=full_rows, mapping=mapping)
    ticker_col_name = mapping.columns['ticker']
    if not ticker_col_name:
        raise Exception("No ticker column defined in mapping.")

    df[ticker_col_name] = tu.normalize_ticker_series(df[ticker_col_name])
    df = tu.filter_ticker_df(df, ticker_col_name, mapping.remove_tickers)
    distinct_tickers = df[ticker_col_name].unique().tolist()    
    return distinct_tickers

def load(etf_name: str | None, file_format: str | None, mapping: Mapping | None, file_name: str | None, raw_data: bytes | None, save: bool = False) -> list[list[str]]:
    try: 

        if file_format == None or mapping == None:
            raise Exception('Missing file type or mapping information in database for data trasformation.')

        if file_name and raw_data:
            if file_format == 'xls':
                if save:
                    with open(os.path.join(FILE_FOLDER, file_name), "wb") as f:
                        f.write(raw_data)
                return read_xls_from_buffer(raw_data, mapping)
            if file_format == 'xlsx':
                if save:
                    with open(os.path.join(FILE_FOLDER, file_name), "wb") as f:
                        f.write(raw_data)
                return read_xlsx_from_buffer(raw_data, mapping)
            elif file_format == 'csv':
                try:
                    str_data = raw_data.decode("utf-8-sig")
                except UnicodeDecodeError:
                    str_data = raw_data.decode("cp1252", errors="replace")

                str_data = str_data.replace("\x00", "")
                if save:
                    with open(os.path.join(FILE_FOLDER, file_name), "w", encoding="utf-8") as f:
                        f.write(str_data)
                return read_csv_from_buffer(str_data, mapping)
            
        raise Exception('Unsupported file format for loading ETF file.')
        
    except Exception as e:
        raise Exception(f"Failed to convert downloaded ETF ({etf_name}) to Data Frame table data: {e}")

